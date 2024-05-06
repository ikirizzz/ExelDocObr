import re, logging
import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font, GradientFill)
from datetime import date

# исходные файла
filename1 = 'gentd.xlsx'
filename2 = 'emp.xlsx'
filename3 = 'spo.xlsx'
# промежуточные стандартизированные
res_filename1 = 'NewQentd.xlsx'
res_filename2 = 'NewEmp.xlsx'
res_filename3 = 'NewSpo.xlsx'
# промежуточный общий
cr_filename1 = 'CRes.xlsx'
# результирующий
pr_filename1 = 'PRIM12.xlsx'

def processing_file(filename, res_filename):
    logging.basicConfig(level=logging.INFO, filename="pylog.log", filemode="a", format="%(asctime)s %(levelname)s %(message)s")
    logging.info("processing_file")

    keywords = ["Ф.И.О. преподавателя", "Кафедра", "Группа"]
    sheet_name = 'очка'
    sheet_name2 = 'спо'
    del_row = 'цикл'
    MAINDATA = {'Ф.И.О. преподавателя': [],
                'Кафедра':              [],
                'цикл':                 [],
                'название дисциплины':  [],
                'Группа':               [],
                'кол-во чел.':         [],
                'кол-во подгр.':       [],
                'сем.':                [],
                'лекц':                [],
                'практ':               [],
                'лаб':                 [],
                'контр. мероприят':    []}
    MAINDATAspo = {'Ф.И.О. преподавателя': [],
                'цикл':                 [],
                'название дисциплины':  [],
                'Группа':               [],
                'кол-во чел.':         [],
                'кол-во подгр.':       [],
                'сем.':                [],
                'лекц':                [],
                'практ':               [],
                'лаб':                 [],
                'контр. мероприят':    []}
    if sheet_name2 in pd.ExcelFile(filename).sheet_names:
        sheet_name = sheet_name2
        MAINDATA = MAINDATAspo

    def find_index_keywords(df, keywords):
        logging.info("start find_index_keywords")
        first_occurrences = {}
        # Поиск только первого вхождения каждого ключевого слова
        for keyword in keywords:
            for column in df.columns:
                # Check if the column contains string values
                if df[column].dtype == 'object':
                    # Используем метод .str.contains() для поиска ключевого слова в столбце
                    matched_rows = df[df[column].astype(str).str.contains(keyword, na=False)]
                    if not matched_rows.empty:
                        row_index = matched_rows.index[0] + 1  # индекс начинается с 0, поэтому добавляем 1
                        first_occurrences[keyword] = [row_index, column]
                        break

        # Если для какого-то ключевого слова не найдено вхождений
        for keyword in keywords:
            if keyword not in first_occurrences:
                logging.warning(f"value '{keyword}' not found.")
        return first_occurrences

    logging.info(f"read_excel '{filename}'")
    df = pd.read_excel(filename, sheet_name= sheet_name)

    results = find_index_keywords(df, keywords)
    row_counter = Counter([result[0] for result in results.values()])
    count = row_counter.most_common(1)[0][0]

    # Определение первой не пустой ячейки в строке count
    first_non_empty_cell = None
    for column in range(len(df.columns)):
        value = df.iloc[count - 1, column]
        if not pd.isna(value):
            first_non_empty_cell = column
            break
    # Начало записи столбцов в MAINDATA с первой не пустой ячейки в строке count
    if first_non_empty_cell is not None:
        for idx, key in enumerate(MAINDATA.keys()):
            column = first_non_empty_cell + idx
            if column < len(df.columns):
                MAINDATA[key] = df.iloc[count - 1:, column].tolist()  # Changed here
            else:
                break
    # Удаление строк, в которых значение "Кадедра" не определено
    undefined_indices = [i for i, val in enumerate(MAINDATA[del_row]) if pd.isna(val)]
    for key in MAINDATA.keys():
        MAINDATA[key] = [val for i, val in enumerate(MAINDATA[key]) if i not in undefined_indices]
    if sheet_name == sheet_name2:
        for key in range (len(MAINDATA["Ф.И.О. преподавателя"])):
            if pd.isna(MAINDATA["Ф.И.О. преподавателя"][key]):
                MAINDATA["Ф.И.О. преподавателя"][key] = MAINDATA["Ф.И.О. преподавателя"][key-1]

    pd.DataFrame(MAINDATA).to_excel(res_filename, sheet_name=sheet_name, index=False)

def creat_file(filename1, filename2, filename3, cr_filename1):
    logging.basicConfig(level=logging.INFO, filename="pylog.log", filemode="a", format="%(asctime)s %(levelname)s %(message)s")
    logging.info("creat_file")

    df1 = pd.read_excel(filename1)
    df2 = pd.read_excel(filename2)
    df3 = pd.read_excel(filename3)
    df = pd.concat([df1, df2, df3], ignore_index=True)

    # Группировка данных по столбцу "Группа"
    grouped = df.groupby("Группа")
    with pd.ExcelWriter(cr_filename1) as writer:
        # Итерация по каждой группе
        for group, data in grouped:
            if group != "Группа":
                Sname = ''.join(filter(str.isdigit, group))
                data.to_excel(writer, sheet_name=f"{Sname}", index=False)

    #проверка на общие предметы групп
    xls = pd.ExcelFile(cr_filename1)
    sheet_names = xls.sheet_names
    for sheet_name in sheet_names:
        if len(sheet_name) > 3:
            split_sheet_names = []
            split_names = [sheet_name[i:i + 3] for i in range(0, len(sheet_name), 3)]
            split_sheet_names.extend(split_names)

            for i in split_sheet_names:
                source_df = pd.read_excel(cr_filename1, sheet_name=sheet_name)
                destination_df = pd.DataFrame()
                if i in xls.sheet_names:
                    destination_df = pd.read_excel(cr_filename1, sheet_name=i)
                destination_df = pd.concat([destination_df, source_df], ignore_index=True)
                with pd.ExcelWriter(cr_filename1, mode='a', engine='openpyxl') as writer:
                    if i in writer.book.sheetnames:
                        idx = writer.book.sheetnames.index(i)
                        writer.book.remove(writer.book.worksheets[idx])
                    destination_df.to_excel(writer, sheet_name=i, index=False)

            wb = load_workbook(cr_filename1)
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            wb.save(cr_filename1)

def copy_data_between_workbooks(source_file, target_file, source_sheet_name, direktor = "В.Н.Борщенюк", napravl = "Направление ", magistr = 0):
    # Загрузить исходную и целевую рабочую книгу и лист
    source_wb = load_workbook(source_file)
    source_sheet = source_wb[source_sheet_name]

    try:
        target_wb = load_workbook(target_file)
        target_sheet = target_wb.create_sheet(source_sheet_name)
    except:
        target_wb = Workbook()
        target_sheet = target_wb.create_sheet(source_sheet_name)
        del target_wb['Sheet']

    # количество строк исходный файл
    filled_cells = 0
    for cell in source_sheet['A']:
        if cell.value is not None:
            filled_cells += 1
    unique_values = list(set(str(cell.value) for cell in source_sheet['H'] if cell.value is not None))
    #не удалая timeSleep он нужен для стабольности, можешь увеличить, но минимум 0.2

    for i in unique_values[::-1]:
        if len(str(i)) > 2:
            unique_values.remove(i)
    if int(unique_values[1]) > int(unique_values[0]): Sem0 = [str(unique_values[0]), str(unique_values[0])+"-"+str(unique_values[1]), str(unique_values[1])]
    else: Sem0 = [str(unique_values[1]), str(unique_values[1])+"-"+str(unique_values[0]), str(unique_values[0])]

    today = date.today()
    Ft = Font(name='Times New Roman', size=8)
    ALNone = Alignment(vertical="bottom", horizontal='left', wrap_text=False)
    AL = Alignment(vertical="center", horizontal='center', wrap_text=True)
    BDNone = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    BDRb = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"),   left=Side(border_style='thin'),   right=Side(border_style='medium'))
    BDRi = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"),   left=Side(border_style='thin'),   right=Side(border_style='thin'))
    BD   = Border(top=Side(border_style="medium"), bottom=Side(border_style="medium"), left=Side(border_style='medium'), right=Side(border_style='medium'))
    BDLLPK  = Border(top=Side(border_style="medium"), bottom=Side(border_style="medium"), left=Side(border_style='thin'), right=Side(border_style='thin'))
    BDLLPK2 = Border(top=Side(border_style="medium"), bottom=Side(border_style="medium"), left=Side(border_style='thin'), right=Side(border_style='medium'))
    def FAB(ICH, V, F, A=ALNone, B = BDNone):
        target_sheet[ICH].value = V
        target_sheet[ICH].font = F
        target_sheet[ICH].alignment = A
        target_sheet[ICH].border = B
    def SumCol(A):
        sum = 0
        for rown in target_sheet.iter_rows(min_row=14, max_row=row-1, min_col=A, max_col=A):
            for cell in rown:
                if cell.value: sum += int(cell.value)
        return sum

    #линейка
    target_sheet.sheet_view.showGridLines = False
    target_sheet.column_dimensions['A'].width = 2.14 + 1.11
    target_sheet.column_dimensions['B'].width = 7.50 + 1.11
    target_sheet.column_dimensions['C'].width = 23.57 + 1.11
    target_sheet.column_dimensions['D'].width = 1.57 + 1.11
    target_sheet.column_dimensions['E'].width = 3.57 + 1.11
    target_sheet.column_dimensions['F'].width = 3.29 + 1.11
    target_sheet.column_dimensions['G'].width = 3.71 + 1.11
    target_sheet.column_dimensions['H'].width = 4.14 + 1.11
    target_sheet.column_dimensions['I'].width = 4.00 + 1.11
    target_sheet.column_dimensions['J'].width = 4.00 + 1.11
    target_sheet.column_dimensions['K'].width = 6.43 + 1.11
    target_sheet.column_dimensions['L'].width = 4.14 + 1.11
    target_sheet.column_dimensions['M'].width = 4.00 + 1.11
    target_sheet.column_dimensions['N'].width = 4.00 + 1.11
    target_sheet.column_dimensions['O'].width = 6.43 + 1.11
    target_sheet.row_dimensions[1].height = 25.75
    target_sheet.row_dimensions[2].height = 15.75
    target_sheet.row_dimensions[3].height = 15.75
    target_sheet.row_dimensions[4].height = 20.75
    target_sheet.row_dimensions[5].height = 20.75
    target_sheet.row_dimensions[6].height = 13.50
    target_sheet.row_dimensions[7].height = 15.75
    target_sheet.row_dimensions[8].height = 15.75
    target_sheet.row_dimensions[9].height = 13.50
    target_sheet.row_dimensions[10].height = 13.50
    target_sheet.row_dimensions[11].height = 11.25
    target_sheet.row_dimensions[12].height = 11.25
    target_sheet.row_dimensions[13].height = 13.50
    target_sheet.merge_cells('A3:G6')
    target_sheet.merge_cells('A7:O7')
    target_sheet.merge_cells('A8:O8')
    target_sheet.merge_cells('A10:A13')
    target_sheet.merge_cells('B10:C13')
    target_sheet.merge_cells('D10:D13')
    target_sheet.merge_cells('E10:G12')
    target_sheet.merge_cells('H10:O10')
    target_sheet.merge_cells('H11:K11')
    target_sheet.merge_cells('H12:K12')
    target_sheet.merge_cells('L11:O11')
    target_sheet.merge_cells('L12:O12')
    #верхний шаблон
    img = Image('IMG.png')
    img.add_image = 'C1'
    target_sheet.add_image(img)
    img.width = 145
    img.height = 60
    img.anchor = 'C1'
    target_sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

    FAB("J1", "УТВЕРЖДАЮ", Font(name='Times New Roman', size=12, bold=True))
    FAB("J2", "Директор",  Font(name='Times New Roman', size=12, bold=False))
    FAB("A3", "МИНИСТЕРСТВО НАУКИ И ВЫСШЕГО                                                   ОБРАЗОВАНИЯ РОССИЙСКОЙ ФЕДЕРАЦИИ                             ФИЛИАЛ ЮЖНО-УРАЛЬСКОГО                              ГОСУДАРСТВЕННОГО УНИВЕРСИТЕТА                                                                                в г. Нижневартовске  ", Font(name='Times New Roman', size=10, bold=False), AL)
    FAB("J3", direktor, Font(name='Times New Roman', size=12, bold=False))
    FAB("J2", "Директор",     Font(name='Times New Roman', size=12, bold=False))
    FAB("O5", f"«____»________________{today.year} г.", Font(name='Times New Roman', size=12), Alignment(vertical="center", horizontal='right'))
    FAB("A7", "РАБОЧИЙ УЧЕБНЫЙ ПЛАН", Font(name='Times New Roman', size=12, bold=True), AL)
    FAB("A8", f"на {today.year}-{int(today.year)+1} учебный год", Font(name='Times New Roman', size=12, bold=True), AL)
    FAB("O9", f"Группа НвФл - {source_sheet_name}", Font(name='Times New Roman', size=10, bold=True), Alignment(vertical="center", horizontal='right'))
    #//////////////////////////
    FAB("A9", "     Направление "+napravl, Font(name='Times New Roman', size=10, bold=True))
    FAB("A10", "№ п/п", Font(name='Times New Roman', size=9), AL, BD)
    FAB("B10", "Название дисциплины",  Font(name='Times New Roman', size=9), AL, BD)
    FAB("D10", "ч/н", Font(name='Times New Roman', size=9), AL, BD)
    FAB("E10", "Объем часов по учебному плану", Font(name='Times New Roman', size=9), AL, BD)
    FAB("H10", "Семестры", Font(name='Times New Roman', size=9), AL, BD)
    FAB("L11", int(Sem0[2]), Font(name='Times New Roman', size=9, bold=True), AL)
    FAB("H11", int(Sem0[0]), Font(name='Times New Roman', size=9, bold=True), AL)

    Sem1 = "16 недель"
    if source_sheet_name[0]!="4": Sem2 = "16 недель"
    else: Sem2 = "16 недель"
    if source_sheet_name=="120": Sem2 = "18 недель"
    if source_sheet_name=="220":
        Sem1 = "12 недель"
        Sem2 = "18 недель"
    FAB("L12", Sem1, Font(name='Times New Roman', size=8), AL)
    FAB("H12", Sem2, Font(name='Times New Roman', size=8), AL)
    FAB("E13", "лекц", Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("F13", "лаб",  Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("G13", "прак", Font(name='Times New Roman', size=8), AL, BDLLPK2)
    FAB("H13", "лекц", Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("I13", "лаб",  Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("J13", "прак", Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("K13", "конт.мер.", Font(name='Times New Roman', size=8), AL, BDLLPK2)
    FAB("L13", "лекц", Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("M13", "лаб",  Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("N13", "прак", Font(name='Times New Roman', size=8), AL, BDLLPK)
    FAB("O13", "конт.мер.", Font(name='Times New Roman', size=8), AL, BDLLPK2)
    for row in target_sheet.iter_rows(min_row=10, max_row=12, min_col=1, max_col=15):
        for cell in row: cell.border = BD
    target_sheet["A13"].border = BD
    target_sheet["B13"].border = BD
    target_sheet["C13"].border = BD
    target_sheet["D13"].border = BD
    #средний шаблон
    row = 14
    Praktiki = []
    for Sem in Sem0:
        for i in range(2, filled_cells+1):
            if str(source_sheet[f"H{i}"].value) == Sem:
                a = 0
                s = 0
                d = 0
                if "Учебная практика" in str(source_sheet[f"D{i}"].value):
                    Praktiki.append(str(source_sheet[f"D{i}"].value))
                    continue
                if "Производственная" in str(source_sheet[f"D{i}"].value):
                    Praktiki.append(str(source_sheet[f"D{i}"].value))
                    continue
                if "ВКР" in str(source_sheet[f"D{i}"].value):
                    Praktiki.append(str(source_sheet[f"D{i}"].value))
                    continue
                if Sem == Sem0[0]:
                    FAB(f"A{row}", row - 12, Font(name='Times New Roman', size=9), AL, BDRb)
                    FAB(f"B{row}", source_sheet[f"C{i}"].value, Font(name='Times New Roman', size=7), AL, BDRb)
                    FAB(f"C{row}", str(source_sheet[f"D{i}"].value) +"\n"+ str(source_sheet[f"A{i}"].value), Font(name='Times New Roman', size=7), Alignment(vertical="center", horizontal='left',   wrap_text=True), BDRb)
                    if source_sheet[f"I{i}"].value: a = int(source_sheet[f"I{i}"].value)
                    if source_sheet[f"J{i}"].value: s = int(source_sheet[f"J{i}"].value)
                    if source_sheet[f"K{i}"].value: d = int(source_sheet[f"K{i}"].value)
                    FAB(f"D{row}", (a+s+d)/16, Ft, AL, BDRb)
                    FAB(f"E{row}", source_sheet[f"I{i}"].value, Ft, AL, BDRi)
                    FAB(f"F{row}", source_sheet[f"K{i}"].value, Ft, AL, BDRi)
                    FAB(f"G{row}", source_sheet[f"J{i}"].value, Ft, AL, BDRb)
                    FAB(f"H{row}", source_sheet[f"I{i}"].value, Ft, AL, BDRi)
                    FAB(f"I{row}", source_sheet[f"K{i}"].value, Ft, AL, BDRi)
                    FAB(f"J{row}", source_sheet[f"J{i}"].value, Ft, AL, BDRi)
                    FAB(f"K{row}", source_sheet[f"L{i}"].value.replace(",", "\n"), Ft, AL, BDRb)
                    FAB(f"L{row}", "", Ft, AL, BDRi)
                    FAB(f"M{row}", "", Ft, AL, BDRi)
                    FAB(f"N{row}", "", Ft, AL, BDRi)
                    FAB(f"O{row}", "", Ft, AL, BDRb)
                if Sem == Sem0[2]:
                    FAB(f"A{row}", row - 12, Font(name='Times New Roman', size=9), AL, BDRb)
                    FAB(f"B{row}", source_sheet[f"C{i}"].value, Font(name='Times New Roman', size=7), AL, BDRb)
                    FAB(f"C{row}", str(source_sheet[f"D{i}"].value) + "\n" + str(source_sheet[f"A{i}"].value), Font(name='Times New Roman', size=7), Alignment(vertical="center", horizontal='left', wrap_text=True), BDRb)
                    if source_sheet[f"I{i}"].value: a = int(source_sheet[f"I{i}"].value)
                    if source_sheet[f"J{i}"].value: s = int(source_sheet[f"J{i}"].value)
                    if source_sheet[f"K{i}"].value: d = int(source_sheet[f"K{i}"].value)
                    FAB(f"D{row}", (a + s + d) / 16, Ft, AL, BDRb)
                    FAB(f"E{row}", source_sheet[f"I{i}"].value, Ft, AL, BDRi)
                    FAB(f"F{row}", source_sheet[f"K{i}"].value, Ft, AL, BDRi)
                    FAB(f"G{row}", source_sheet[f"J{i}"].value, Ft, AL, BDRb)
                    FAB(f"H{row}", "", Ft, AL, BDRi)
                    FAB(f"I{row}", "", Ft, AL, BDRi)
                    FAB(f"J{row}", "", Ft, AL, BDRi)
                    FAB(f"K{row}", "", Ft, AL, BDRb)
                    FAB(f"L{row}", source_sheet[f"I{i}"].value, Ft, AL, BDRi)
                    FAB(f"M{row}", source_sheet[f"K{i}"].value, Ft, AL, BDRi)
                    FAB(f"N{row}", source_sheet[f"J{i}"].value, Ft, AL, BDRi)
                    if source_sheet[f"L{i}"].value: FAB(f"O{row}", source_sheet[f"L{i}"].value.replace(",", "\n"), Ft, AL, BDRb)
                if Sem == Sem0[1]:
                    FAB(f"A{row}", row - 12, Font(name='Times New Roman', size=9), AL, BDRb)
                    FAB(f"B{row}", source_sheet[f"C{i}"].value, Font(name='Times New Roman', size=7), AL, BDRb)
                    FAB(f"C{row}", str(source_sheet[f"D{i}"].value) + "\n" + str(source_sheet[f"A{i}"].value), Font(name='Times New Roman', size=7), Alignment(vertical="center", horizontal='left', wrap_text=True), BDRb)
                    FAB(f"E{row}", source_sheet[f"I{i}"].value, Ft, AL, BDRi)
                    FAB(f"F{row}", source_sheet[f"K{i}"].value, Ft, AL, BDRi)
                    FAB(f"G{row}", source_sheet[f"J{i}"].value, Ft, AL, BDRb)
                    if target_sheet[f"E{row}"].value:
                        FAB(f"H{row}", int(target_sheet[f"E{row}"].value)-(int(target_sheet[f"E{row}"].value) // 16)*8, Ft, AL, BDRi)
                        FAB(f"L{row}", (int(target_sheet[f"E{row}"].value) // 16) * 8, Ft, AL, BDRi)
                        a = [target_sheet[f"H{row}"].value, target_sheet[f"L{row}"].value]
                    else:
                        FAB(f"L{row}", "", Ft, AL, BDRi)
                        FAB(f"H{row}", "", Ft, AL, BDRi)
                        a = [0, 0]
                    if target_sheet[f"F{row}"].value:
                        FAB(f"I{row}", int(target_sheet[f"F{row}"].value)-(int(target_sheet[f"F{row}"].value) // 16)*8, Ft, AL, BDRi)
                        FAB(f"M{row}", (int(target_sheet[f"F{row}"].value) // 16) * 8, Ft, AL, BDRi)
                        s = [target_sheet[f"I{row}"].value, target_sheet[f"M{row}"].value]
                    else:
                        FAB(f"I{row}", "", Ft, AL, BDRi)
                        FAB(f"M{row}", "", Ft, AL, BDRi)
                        s = [0, 0]
                    if target_sheet[f"G{row}"].value:
                        FAB(f"J{row}", int(target_sheet[f"G{row}"].value)-(int(target_sheet[f"G{row}"].value) // 16)*8, Ft, AL, BDRi)
                        FAB(f"N{row}", (int(target_sheet[f"G{row}"].value) // 16) * 8, Ft, AL, BDRi)
                        d = [target_sheet[f"J{row}"].value, target_sheet[f"N{row}"].value]
                    else:
                        FAB(f"J{row}", "", Ft, AL, BDRi)
                        FAB(f"N{row}", "", Ft, AL, BDRi)
                        d = [0,0]

                    sp1 = ""
                    sp2 = ""
                    Fho = 0
                    spliti = re.split(r'[^а-яА-я^a-zA-Z0-9.]', source_sheet[f"L{i}"].value)
                    spliti = list(filter(None, spliti))
                    for ho in range(len(spliti)-1, -1, -1):
                        if "ф" in spliti[ho]:
                            spliti[ho] = "диф.зач."
                            del spliti[ho + 1]
                        elif "эк" in spliti[ho]:
                            spliti[ho] = "экзамен"
                        elif "за" in spliti[ho]:
                            spliti[ho] = "зачет"
                    for ho in range(len(spliti)):
                        if Fho == 1:
                            Fho =0
                            continue
                        if "2" in spliti[ho]:
                            sp1 += " "+ spliti[ho + 1]
                            sp2 += " "+ spliti[ho + 1]
                            Fho =1
                            continue
                        if sp1 == "": sp1 += " "+ spliti[ho]
                        elif (sp1 != "") and (sp2 == "") and ("з" not in spliti[ho]) and ("ф" not in spliti[ho]) and ("э" not in spliti[ho]):  sp1 += " "+ spliti[ho]
                        else: sp2 += " "+ spliti[ho]

                    FAB(f"K{row}", sp1, Ft, AL, BDRb)
                    FAB(f"O{row}", sp2, Ft, AL, BDRb)
                    FAB(f"D{row}", f"{int((a[0] + s[0] + d[0]) / 16)} \n{int((a[1] + s[1] + d[1]) / 16)}" , Ft, AL, BDRb)
                rd = target_sheet.row_dimensions[row]  # выбор ряда
                if source_sheet[f"D{i}"].value and len(source_sheet[f"D{i}"].value) < 36:
                    rd.height = 23  # высота ряда
                elif source_sheet[f"D{i}"].value and len(source_sheet[f"D{i}"].value) > 70:
                    rd.height = 44
                else: rd.height = 33
                row+=1
    #нижний шаблон
    if magistr == 0:
        target_sheet.row_dimensions[row].height = 38.5
        target_sheet.row_dimensions[row+1].height = 13.5
        target_sheet.row_dimensions[row+2].height = 17.0
        target_sheet.row_dimensions[row+3].height = 17.0
        target_sheet.row_dimensions[row+4].height = 22.5
        target_sheet.row_dimensions[row+5].height = 22.5
        target_sheet.row_dimensions[row+6].height = 22.5
        target_sheet.merge_cells(f'A{row}:D{row}')
        target_sheet.merge_cells(f'A{row+1}:G{row+1}')
        target_sheet.merge_cells(f'H{row+1}:K{row+1}')
        target_sheet.merge_cells(f'L{row+1}:O{row+1}')
        target_sheet.merge_cells(f'A{row+2}:B{row+3}')
        target_sheet.merge_cells(f'C{row+2}:D{row+2}')
        target_sheet.merge_cells(f'E{row+2}:G{row+2}')
        target_sheet.merge_cells(f'H{row+2}:K{row+2}')
        target_sheet.merge_cells(f'L{row+2}:O{row+2}')
        target_sheet.merge_cells(f'C{row+3}:D{row+3}')
        target_sheet.merge_cells(f'E{row+3}:G{row+3}')
        target_sheet.merge_cells(f'H{row+3}:K{row+3}')
        target_sheet.merge_cells(f'L{row+3}:O{row+3}')
        FAB(f"A{row}", "Итого учебных занятий                                                                  (без подгрупп) ", Font(name='Times New Roman', size=9), AL)
        FAB(f"A{row+1}", "Итого в неделю без физической культуры", Font(name='Times New Roman', size=9), AL)
        FAB(f"A{row+2}", "Зачетно - экзаменационная сессия", Font(name='Times New Roman', size=7), AL)
        FAB(f"C{row+2}", "Зачетная ", Font(name='Times New Roman', size=9), AL)
        FAB(f"C{row+3}", "Экзаменационная ", Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row+2}", "2 недели", Font(name='Times New Roman', size=9), AL)
        week = "5 недель"
        DatZ1 =f"23.12.{today.year%100-1}-30.12.{today.year%100-1}"
        DatS1 =f"09.01.{today.year%100}-29.01.{today.year%100}"
        DatZ2 =f"05.06.{today.year%100}-11.06.{today.year%100}"
        DatS2 =f"12.06.{today.year%100}-25.06.{today.year%100}"
        if source_sheet_name[0] != "4":
            DatS1 = f"09.01.{today.year % 100}-22.01.{today.year % 100}"
            DatZ2 = f"15.05.{today.year % 100}-21.05.{today.year % 100}"
            DatS2 = f"22.05.{today.year % 100}-04.06.{today.year % 100}"
        if source_sheet_name[1]+source_sheet_name[2] == "20":
            DatZ1 = f"14.11.{today.year % 100 - 1}-20.11.{today.year % 100 - 1}"
            DatS1 = f"19.12.{today.year % 100 - 1}-25.12.{today.year % 100 - 1}"
            DatZ2 = f"08.05.{today.year % 100}-14.05.{today.year % 100}"
            DatS2 = f"26.06.{today.year % 100}-02.07.{today.year % 100}"
        FAB(f"H{row+2}", DatZ1, Font(name='Times New Roman', size=9), AL)
        FAB(f"H{row+3}", DatS1, Font(name='Times New Roman', size=9), AL)
        FAB(f"L{row+2}", DatZ2, Font(name='Times New Roman', size=9), AL)
        FAB(f"L{row+3}", DatS2, Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row+3}", week, Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row}", SumCol(5), Ft, AL)
        FAB(f"F{row}", SumCol(6), Ft, AL)
        FAB(f"G{row}", SumCol(7), Ft, AL)
        FAB(f"H{row}", SumCol(8), Ft, AL)
        FAB(f"I{row}", SumCol(9), Ft, AL)
        FAB(f"J{row}", SumCol(10), Ft, AL)
        FAB(f"L{row}", SumCol(12), Ft, AL)
        FAB(f"M{row}", SumCol(13), Ft, AL)
        FAB(f"N{row}", SumCol(14), Ft, AL)

        spliti = ""
        spliti0 = ""
        for rown in target_sheet.iter_rows(min_row=14, max_row=row - 1, min_col=11, max_col=11):
            for cell in rown:
                if cell.value:
                    spliti += cell.value
        if spliti.count('э')>0: spliti0 += f"{spliti.count('э')} экзам.\n"
        if spliti.count('ди')>0: spliti0 += f"{spliti.count('ди')} д.зач.\n"
        if spliti.count('зач')-spliti.count('ди')>0: spliti0 += f"{spliti.count('зач')-spliti.count('ди')} зач.\n"
        if spliti.count('кур')-spliti.count('пр')>0: spliti0 += f"{spliti.count('кур')-spliti.count('пр')} кур.р.\n"
        if spliti.count('пр')>0: spliti0 += f"{spliti.count('пр')} курс.пр."
        FAB(f"K{row}", spliti0[:-2], Font(name='Times New Roman', size=7), AL)
        spliti = ""
        spliti0 = ""
        for rown in target_sheet.iter_rows(min_row=14, max_row=row - 1, min_col=11, max_col=11):
            for cell in rown:
                if cell.value:
                    spliti += cell.value
        if spliti.count('э')>0: spliti0 += f"{spliti.count('э')} экзам.\n"
        if spliti.count('ди')>0: spliti0 += f"{spliti.count('ди')} д.зач.\n"
        if spliti.count('зач')-spliti.count('ди')>0: spliti0 += f"{spliti.count('зач')-spliti.count('ди')} зач.\n"
        if spliti.count('кур')-spliti.count('пр')>0: spliti0 += f"{spliti.count('кур')-spliti.count('пр')} кур.р.\n"
        if spliti.count('пр')>0: spliti0 += f"{spliti.count('пр')} курс.пр."
        FAB(f"O{row}", spliti0[:-2], Font(name='Times New Roman', size=7), AL)

        fleg =0
        fleg2 = 0
        for rown in target_sheet.iter_rows(min_row=14, max_row=row-1, min_col=3, max_col=3):
            for cell in rown:
                if "Физическая культура" in str(cell.value):
                    if target_sheet[f"H{cell.row}"].value and target_sheet[f"H{cell.row}"].value != "": fleg += int(target_sheet[f"H{cell.row}"].value)
                    if target_sheet[f"I{cell.row}"].value and target_sheet[f"I{cell.row}"].value != "": fleg += int(target_sheet[f"I{cell.row}"].value)
                    if target_sheet[f"J{cell.row}"].value and target_sheet[f"J{cell.row}"].value != "": fleg += int(target_sheet[f"J{cell.row}"].value)
                    if target_sheet[f"L{cell.row}"].value and target_sheet[f"L{cell.row}"].value != "": fleg2 += int(target_sheet[f"L{cell.row}"].value)
                    if target_sheet[f"M{cell.row}"].value and target_sheet[f"M{cell.row}"].value != "": fleg2 += int(target_sheet[f"M{cell.row}"].value)
                    if target_sheet[f"N{cell.row}"].value and target_sheet[f"N{cell.row}"].value != "": fleg2 += int(target_sheet[f"N{cell.row}"].value)
                    break
        if fleg == 0:
            FAB(f"H{row + 1}", float(SumCol(8) + SumCol(9) + SumCol(10)) / 16, Ft, AL)
            FAB(f"L{row + 1}", float(SumCol(12) + SumCol(13) + SumCol(14)) / 16, Ft, AL)
        else:
            FAB(f"H{row + 1}", float(SumCol(8) + SumCol(9) + SumCol(10) - fleg) / 16, Ft, AL)
            FAB(f"L{row + 1}", float(SumCol(12) + SumCol(13) + SumCol(14) - fleg2) / 16, Ft, AL)
        Praktiki = sorted(Praktiki, reverse=True)
        lenP = len(Praktiki)
        for ind in range(lenP):
            target_sheet.merge_cells(f'A{row+4+ind}:D{row+4+ind}')
            target_sheet.merge_cells(f'E{row+4+ind}:G{row+4+ind}')
            target_sheet.merge_cells(f'H{row+4+ind}:O{row+4+ind}')
            FAB(f"A{row+4+ind}", Praktiki[ind], Ft, AL)
            if "ВКР" in Praktiki[ind]: FAB(f"E{row+4+ind}", "6 недель", Ft, AL)
            elif "Учебная практика" in Praktiki[ind]:
                FAB(f"E{row+4+ind}", f"{Praktiki[ind].count('Учебная')*2} недели", Ft, AL)
            else: FAB(f"E{row+4+ind}", "4 недели", Ft, AL)
        target_sheet.merge_cells(f'A{row+4+lenP}:D{row+4+lenP}')
        target_sheet.merge_cells(f'E{row+4+lenP}:G{row+4+lenP}')
        FAB(f"A{row+4+lenP}", "Каникулы", Ft, AL)
        target_sheet.merge_cells(f'E{row+4+lenP}:G{row+4+lenP}')
        target_sheet.merge_cells(f'H{row+4+lenP}:K{row+4+lenP}')
        target_sheet.merge_cells(f'L{row+4+lenP}:O{row+4+lenP}')

        for rown in target_sheet.iter_rows(min_row=row, max_row=row+4+lenP, min_col=1, max_col=15):
            for cell in rown: cell.border = BD
        for rown in target_sheet.iter_rows(min_row=14, max_row=row, min_col=15, max_col=15):
            for cell in rown: cell.border = BDRb
    if magistr == 1:
        target_sheet.row_dimensions[row].height = 38.5
        target_sheet.row_dimensions[row + 1].height = 13.5
        target_sheet.row_dimensions[row + 2].height = 22.0
        target_sheet.row_dimensions[row + 3].height = 22.0
        target_sheet.row_dimensions[row + 4].height = 22.0
        target_sheet.row_dimensions[row + 5].height = 33.0
        target_sheet.row_dimensions[row + 6].height = 33.0
        target_sheet.merge_cells(f'A{row}:D{row}')
        target_sheet.merge_cells(f'A{row + 1}:G{row + 1}')
        target_sheet.merge_cells(f'H{row + 1}:K{row + 1}')
        target_sheet.merge_cells(f'L{row + 1}:O{row + 1}')
        target_sheet.merge_cells(f'A{row + 2}:B{row + 4}')
        target_sheet.merge_cells(f'C{row + 2}:D{row + 2}')
        target_sheet.merge_cells(f'E{row + 2}:G{row + 2}')
        target_sheet.merge_cells(f'H{row + 2}:K{row + 2}')
        target_sheet.merge_cells(f'L{row + 2}:O{row + 2}')
        target_sheet.merge_cells(f'C{row + 3}:D{row + 3}')
        target_sheet.merge_cells(f'E{row + 3}:G{row + 3}')
        target_sheet.merge_cells(f'C{row + 4}:D{row + 4}')
        target_sheet.merge_cells(f'E{row + 4}:G{row + 4}')
        FAB(f"A{row}",
            "Итого учебных занятий                                                                  (без подгрупп) ",
            Font(name='Times New Roman', size=9), AL)
        FAB(f"A{row + 1}", "Итого аудиторной нагрузки  (без подгрупп)", Font(name='Times New Roman', size=9), AL)
        FAB(f"A{row + 2}", "Cроки сессии", Font(name='Times New Roman', size=7), AL)
        FAB(f"C{row + 2}", "", Font(name='Times New Roman', size=9), AL)
        FAB(f"C{row + 3}", "", Font(name='Times New Roman', size=9), AL)
        FAB(f"C{row + 4}", "", Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row + 2}", " деней", Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row + 3}", " деней", Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row + 4}", " деней", Font(name='Times New Roman', size=9), AL)
        FAB(f"H{row + 2}", "средняя нагрузка на день", Font(name='Times New Roman', size=9), AL)
        FAB(f"E{row}", SumCol(5), Ft, AL)
        FAB(f"F{row}", SumCol(6), Ft, AL)
        FAB(f"G{row}", SumCol(7), Ft, AL)
        FAB(f"H{row}", SumCol(8), Ft, AL)
        FAB(f"I{row}", SumCol(9), Ft, AL)
        FAB(f"J{row}", SumCol(10), Ft, AL)
        FAB(f"L{row}", SumCol(12), Ft, AL)
        FAB(f"M{row}", SumCol(13), Ft, AL)
        FAB(f"N{row}", SumCol(14), Ft, AL)

        spliti = ""
        spliti0 = ""
        for rown in target_sheet.iter_rows(min_row=14, max_row=row - 1, min_col=11, max_col=11):
            for cell in rown:
                if cell.value:
                    spliti += cell.value
        if spliti.count('э') > 0: spliti0 += f"{spliti.count('э')} экзам.\n"
        if spliti.count('ди') > 0: spliti0 += f"{spliti.count('ди')} д.зач.\n"
        if spliti.count('зач') - spliti.count('ди') > 0: spliti0 += f"{spliti.count('зач') - spliti.count('ди')} зач.\n"
        if spliti.count('кур') - spliti.count('пр') > 0: spliti0 += f"{spliti.count('кур') - spliti.count('пр')} кур.р.\n"
        if spliti.count('пр') > 0: spliti0 += f"{spliti.count('пр')} курс.пр."
        FAB(f"K{row}", spliti0[:-2], Font(name='Times New Roman', size=7), AL)
        spliti = ""
        spliti0 = ""
        for rown in target_sheet.iter_rows(min_row=14, max_row=row - 1, min_col=11, max_col=11):
            for cell in rown:
                if cell.value:
                    spliti += cell.value
        if spliti.count('э') > 0: spliti0 += f"{spliti.count('э')} экзам.\n"
        if spliti.count('ди') > 0: spliti0 += f"{spliti.count('ди')} д.зач.\n"
        if spliti.count('зач') - spliti.count('ди') > 0: spliti0 += f"{spliti.count('зач') - spliti.count('ди')} зач.\n"
        if spliti.count('кур') - spliti.count('пр') > 0: spliti0 += f"{spliti.count('кур') - spliti.count('пр')} кур.р.\n"
        if spliti.count('пр') > 0: spliti0 += f"{spliti.count('пр')} курс.пр."
        FAB(f"O{row}", spliti0[:-2], Font(name='Times New Roman', size=7), AL)

        FAB(f"H{row + 1}", float(SumCol(8) + SumCol(9) + SumCol(10)), Ft, AL)
        FAB(f"L{row + 1}", float(SumCol(12) + SumCol(13) + SumCol(14)), Ft, AL)

        if source_sheet_name[0] == "5":
            target_sheet.merge_cells(f'A{row + 5}:D{row + 5}')
            target_sheet.merge_cells(f'E{row + 5}:G{row + 5}')
            target_sheet.merge_cells(f'A{row + 6}:D{row + 6}')
            target_sheet.merge_cells(f'E{row + 6}:G{row + 6}')
            target_sheet.merge_cells(f'H{row + 3}:K{row + 6}')
            target_sheet.merge_cells(f'L{row + 3}:O{row + 6}')
            FAB(f"H{row + 3}", "Производственная (преддипломная) практика (4 недели)", Font(name='Times New Roman', size=7), AL)
            FAB(f"A{row + 5}", "Подготовка к сдаче и сдача государственного экзамена (2 недели)", Font(name='Times New Roman', size=7), AL)
            FAB(f"A{row + 6}", "Защита ВКР, включая подготовку к процедуре защиты и процедуру защиты (4 недели)", Font(name='Times New Roman', size=7), AL)
            for rown in target_sheet.iter_rows(min_row=row, max_row=row + 6, min_col=1, max_col=15):
                for cell in rown: cell.border = BD
        else:
            target_sheet.merge_cells(f'H{row + 3}:K{row + 4}')
            target_sheet.merge_cells(f'L{row + 3}:O{row + 4}')
            FAB(f"H{row+3}", max(Praktiki), Font(name='Times New Roman', size=7), AL)
            for rown in target_sheet.iter_rows(min_row=row, max_row=row + 4, min_col=1, max_col=15):
                for cell in rown: cell.border = BD

        for rown in target_sheet.iter_rows(min_row=14, max_row=row-1, min_col=15, max_col=15):
            for cell in rown: cell.border = BDRb
    # Сохраните изменения в целевой рабочей книге
    target_wb.save(target_file)


#есть нагрузки (3 штуки) мы их подгружаем, чтобы стандартизировать (создаются новые файлы)
processing_file(filename1, res_filename1)
processing_file(filename2, res_filename2)
processing_file(filename3, res_filename3)

#создаем общий файл из предыдущих, он также обработывается и собирает все предметы кафедр группам
creat_file(res_filename1, res_filename2, res_filename3, cr_filename1)

# Вызовите функцию для копирования данных
copy_data_between_workbooks(cr_filename1, pr_filename1, '308', magistr=1, napravl= "Экономика")
copy_data_between_workbooks(cr_filename1, pr_filename1, '508', magistr=1, napravl= "Экономика")
copy_data_between_workbooks(cr_filename1, pr_filename1, '129', napravl= "Строительство")
copy_data_between_workbooks(cr_filename1, pr_filename1, '229', napravl= "Строительство")
copy_data_between_workbooks(cr_filename1, pr_filename1, '329', napravl= "Строительство")
copy_data_between_workbooks(cr_filename1, pr_filename1, '429', napravl= "Строительство")